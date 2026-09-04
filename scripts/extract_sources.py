import csv
import json
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import load_workbook

WORKBOOK = Path('/home/ubuntu/Downloads/Controle de Viagens_Veículos.xlsx')
MAIN_CSV = Path('/home/ubuntu/Downloads/BARRA DO GARÇAS - Respostas ao formulário 1.csv')
OUTPUT = Path('/home/ubuntu/controle-viagens/scripts/import_payload.json')
MONTHS = {
    'JANEIRO': 1, 'FEVEREIRO': 2, 'MARÇO': 3, 'ABRIL': 4, 'MAIO': 5, 'JUNHO': 6,
    'JULHO': 7, 'AGOSTO': 8, 'SETEMBRO': 9, 'OUTUBRO': 10, 'NOVEMBRO': 11, 'DEZEMBRO': 12,
}

def clean(value):
    return re.sub(r'\s+', ' ', str(value or '').replace('\n', ' ').strip())

def normalize(value):
    text = unicodedata.normalize('NFD', clean(value)).encode('ascii', 'ignore').decode().upper()
    return re.sub(r'[^A-Z0-9]+', '', text)

def vehicle_display(value):
    text = clean(value).upper()
    text = re.sub(r'\s*\([^)]*\)', '', text)
    text = re.sub(r'\s*[-–—]\s*', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip(' -')
    return text

def vehicle_key(value):
    text = vehicle_display(value)
    return normalize(text)

def number(value):
    text = clean(value).replace('.', '').replace(',', '.')
    text = re.sub(r'[^0-9.-]', '', text)
    try:
        return int(round(float(text))) if text else 0
    except ValueError:
        return 0

def iso_date(day, month, year):
    return datetime(year, month, int(day or 1), 12, 0, 0, tzinfo=timezone.utc).isoformat().replace('+00:00', 'Z')

def parse_main_date(value):
    match = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})(?:\s+(\d{1,2}):(\d{2}):(\d{2}))?', clean(value))
    if not match:
        return None
    return datetime(int(match.group(3)), int(match.group(2)), int(match.group(1)), int(match.group(4) or 12), int(match.group(5) or 0), int(match.group(6) or 0), tzinfo=timezone.utc).isoformat().replace('+00:00', 'Z')

def nonempty(value):
    return clean(value) not in ('', '-', '—')

monthly_trips = []
vehicles = {}
wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
for sheet_name in wb.sheetnames:
    match = re.match(r'^(JANEIRO|FEVEREIRO|MARÇO|ABRIL|MAIO|JUNHO|JULHO|AGOSTO|SETEMBRO|OUTUBRO|NOVEMBRO|DEZEMBRO)\s+(2022|2023|2024|2025|2026)$', sheet_name.upper())
    if not match:
        continue
    month_name, year_text = match.groups()
    year = int(year_text)
    month = MONTHS[month_name]
    if (year, month) < (2022, 5) or (year, month) > (2026, 8):
        continue
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, ())
    vehicle_header_row = next(rows, ())
    headers = [clean(value) for value in vehicle_header_row]
    current_day = 1
    for index, name in enumerate(headers[7:], start=7):
        if nonempty(name):
            display = vehicle_display(name)
            vehicles.setdefault(vehicle_key(display), {'plate': display, 'model': display.split(' ')[0] if display else None, 'category': 'Frota operacional', 'year': year})
    for raw in rows:
        cells = list(raw)
        while len(cells) < max(len(header), len(headers), 13):
            cells.append(None)
        if cells[0] is not None and str(cells[0]).strip() not in ('', 'X', 'x'):
            try:
                current_day = int(float(cells[0]))
            except (TypeError, ValueError):
                pass
        turn = clean(cells[2])
        destination = clean(cells[3])
        action = clean(cells[4])
        driver = clean(cells[5])
        passengers = clean(cells[6])
        markers = []
        for index, name in enumerate(headers[7:], start=7):
            marker = clean(cells[index] if index < len(cells) else '')
            if name and marker.lower() in {'x', 'sim', '1', 'true'}:
                display = vehicle_display(name)
                markers.append(display)
        if not any([turn, destination, action, driver, passengers, markers]):
            continue
        if not markers:
            markers = ['NÃO INFORMADO']
        notes = clean(cells[12] if len(cells) > 12 else '') or None
        for vehicle in markers:
            key = vehicle_key(vehicle)
            if key and key != vehicle_key('NÃO INFORMADO'):
                vehicles.setdefault(key, {'plate': vehicle, 'model': vehicle.split(' ')[0] if vehicle else None, 'category': 'Frota operacional', 'year': year})
            monthly_trips.append({
                'tripDate': iso_date(current_day, month, year),
                'vehiclePlate': vehicle,
                'vehicleModel': vehicle.split(' ')[0] if vehicle != 'NÃO INFORMADO' else None,
                'driverName': driver or 'Não informado',
                'origin': 'Barra do Garças, MT',
                'destination': destination or None,
                'purpose': action or 'Controle de viagem',
                'distanceKm': 0,
                'durationMinutes': 0,
                'status': 'Concluída',
                'notes': 'Turno: ' + turn + ((' | Passageiros: ' + passengers) if passengers else '') + ((' | ' + notes) if notes else ''),
                'importedFile': WORKBOOK.name,
                'sourceSheet': sheet_name,
            })

records = []
with MAIN_CSV.open('r', encoding='utf-8-sig', newline='') as handle:
    reader = csv.reader(handle)
    header = next(reader, [])
    for row in reader:
        row = list(row)
        while len(row) < 21:
            row.append('')
        recorded_at = parse_main_date(row[0])
        if not recorded_at:
            continue
        vehicle = vehicle_display(row[3]) or 'NÃO INFORMADO'
        key = vehicle_key(vehicle)
        vehicles.setdefault(key, {'plate': vehicle, 'model': vehicle.split(' ')[0] if vehicle else None, 'category': 'Frota operacional', 'year': None})
        irregularities = ' | '.join(value for value in [clean(row[9]), clean(row[16])] if nonempty(value)) or None
        fuel = ' / '.join(value for value in [clean(row[12]), clean(row[13])] if nonempty(value)) or None
        condition = ' / '.join(value for value in [clean(row[14]), clean(row[15])] if nonempty(value)) or None
        records.append({
            'recordedAt': recorded_at,
            'respondentName': clean(row[1]) or None,
            'employeeId': clean(row[2]) or None,
            'vehiclePlate': vehicle,
            'event': clean(row[4]) or None,
            'kmInitial': number(row[5]),
            'kmFinal': number(row[6]),
            'serviceType': clean(row[7]) or None,
            'summary': clean(row[8]) or None,
            'fuelLevel': fuel,
            'vehicleCondition': condition,
            'irregularity': irregularities,
            'email': clean(row[17]) or None,
            'declaration': clean(row[20]) or None,
            'importedFile': MAIN_CSV.name,
        })

# Canonicalize every trip/record reference to the same deduplicated vehicle label.
vehicle_by_key = {key: value['plate'] for key, value in vehicles.items()}
for trip in monthly_trips:
    trip['vehiclePlate'] = vehicle_by_key.get(vehicle_key(trip['vehiclePlate']), trip['vehiclePlate'])
for record in records:
    record['vehiclePlate'] = vehicle_by_key.get(vehicle_key(record['vehiclePlate']), record['vehiclePlate'])

payload = {'monthlyTrips': monthly_trips, 'records': records, 'vehicles': list(vehicles.values()), 'monthlySheets': sorted({trip['sourceSheet'] for trip in monthly_trips})}
OUTPUT.write_text(json.dumps(payload, ensure_ascii=False), encoding='utf-8')
print(json.dumps({'monthlySheets': len(payload['monthlySheets']), 'monthlyTrips': len(monthly_trips), 'records': len(records), 'vehicles': len(payload['vehicles']), 'sheets': payload['monthlySheets']}, ensure_ascii=False))
