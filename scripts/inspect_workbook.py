from pathlib import Path
from openpyxl import load_workbook
file = Path('/home/ubuntu/Downloads/Controle de Viagens_Veículos.xlsx')
wb = load_workbook(file, read_only=True, data_only=True)
print('total_sheets=', len(wb.sheetnames))
for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.iter_rows(min_row=1, max_row=4, values_only=True))
    print(name, 'rows=', ws.max_row, 'cols=', ws.max_column, 'sample=', rows)
